"""
Excel 导出模块
生成专业格式的测试用例 Excel 文件
"""
import io
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# 样式定义
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_FONT = Font(name="微软雅黑", size=10)
CELL_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

# 优先级颜色
PRIORITY_COLORS = {
    "P0": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
    "P1": PatternFill(start_color="FF8800", end_color="FF8800", fill_type="solid"),
    "P2": PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid"),
    "P3": PatternFill(start_color="88CC88", end_color="88CC88", fill_type="solid"),
}

PRIORITY_FONTS = {
    "P0": Font(name="微软雅黑", size=10, bold=True, color="FFFFFF"),
    "P1": Font(name="微软雅黑", size=10, bold=True, color="FFFFFF"),
    "P2": Font(name="微软雅黑", size=10, bold=True),
    "P3": Font(name="微软雅黑", size=10),
}

# 交替行颜色
EVEN_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

# 表头定义
HEADERS = [
    ("用例编号", 12),
    ("所属模块", 18),
    ("用例标题", 35),
    ("优先级", 8),
    ("用例类型", 12),
    ("前置条件", 25),
    ("测试步骤", 45),
    ("预期结果", 40),
]


def export_to_excel(test_cases: List[Dict[str, Any]], title: str = "测试用例") -> bytes:
    """
    将测试用例导出为 Excel 文件

    参数:
        test_cases: 测试用例列表
        title: 工作表标题

    返回:
        Excel 文件的 bytes 数据
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # 设置列宽
    for col_idx, (_, width) in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 写入标题行（合并单元格）
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title_cell = ws.cell(row=1, column=1, value=f"📋 {title}")
    title_cell.font = Font(name="微软雅黑", size=14, bold=True, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # 写入统计行
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    p0_count = sum(1 for tc in test_cases if tc.get("priority") == "P0")
    p1_count = sum(1 for tc in test_cases if tc.get("priority") == "P1")
    p2_count = sum(1 for tc in test_cases if tc.get("priority") == "P2")
    p3_count = sum(1 for tc in test_cases if tc.get("priority") == "P3")
    stats_cell = ws.cell(row=2, column=1,
        value=f"总计: {len(test_cases)} 条 | P0: {p0_count} | P1: {p1_count} | P2: {p2_count} | P3: {p3_count}")
    stats_cell.font = Font(name="微软雅黑", size=10, color="666666")
    stats_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    # 写入表头
    header_row = 3
    for col_idx, (header_name, _) in enumerate(HEADERS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 28

    # 写入数据行
    for row_idx, tc in enumerate(test_cases, 1):
        data_row = header_row + row_idx
        is_even = row_idx % 2 == 0

        values = [
            tc.get("case_id", ""),
            tc.get("module", ""),
            tc.get("title", ""),
            tc.get("priority", "P2"),
            tc.get("case_type", ""),
            tc.get("precondition", ""),
            tc.get("steps", ""),
            tc.get("expected_result", ""),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=data_row, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT if col_idx > 2 else CENTER_ALIGNMENT
            cell.border = THIN_BORDER

            # 交替行颜色
            if is_even:
                cell.fill = EVEN_ROW_FILL

        # 优先级特殊样式
        priority = tc.get("priority", "P2")
        priority_cell = ws.cell(row=data_row, column=4)
        if priority in PRIORITY_COLORS:
            priority_cell.fill = PRIORITY_COLORS[priority]
            priority_cell.font = PRIORITY_FONTS.get(priority, CELL_FONT)
            priority_cell.alignment = CENTER_ALIGNMENT

        # 设置行高
        ws.row_dimensions[data_row].height = max(40, 20 * max(
            values[6].count("\n") + 1,  # steps 行数
            values[7].count("\n") + 1,  # expected_result 行数
            2
        ))

    # 冻结窗格（冻结表头）
    ws.freeze_panes = f"A{header_row + 1}"

    # 自动筛选
    ws.auto_filter.ref = f"A{header_row}:H{header_row + len(test_cases)}"

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

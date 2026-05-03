"""
Excel/CSV 文件解析器
支持 .xlsx, .xls, .csv 格式的测试用例导入
"""


def parse_excel(file_path: str) -> list:
    """Parse Excel file and return list of test case dicts"""
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(max_row=1))]

    # Map common header names to our fields
    field_map = {
        'title': 'title', '用例标题': 'title', '标题': 'title', '名称': 'title',
        'module': 'module', '模块': 'module', '功能模块': 'module',
        'steps': 'steps', '测试步骤': 'steps', '步骤': 'steps', '操作步骤': 'steps',
        'expected_result': 'expected_result', '预期结果': 'expected_result', '期望结果': 'expected_result',
        'priority': 'priority', '优先级': 'priority',
        'case_type': 'case_type', '用例类型': 'case_type', '类型': 'case_type',
        'precondition': 'precondition', '前置条件': 'precondition',
        'case_id': 'case_id', '用例编号': 'case_id', '编号': 'case_id',
    }

    col_map = {}
    for i, h in enumerate(headers):
        if h in field_map:
            col_map[i] = field_map[h]

    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        tc = {'title': '', 'steps': '', 'expected_result': '', 'priority': 'P2', 'case_type': '功能测试', 'module': '', 'precondition': '', 'case_id': ''}
        for i, val in enumerate(row):
            if i in col_map and val:
                tc[col_map[i]] = str(val).strip()
        if tc['title']:  # skip empty rows
            cases.append(tc)
    wb.close()
    return cases


def parse_csv_import(file_path: str) -> list:
    """Parse CSV file and return list of test case dicts"""
    import csv
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        # Same field mapping
        field_map = {
            'title': 'title', '用例标题': 'title', '标题': 'title', '名称': 'title',
            'module': 'module', '模块': 'module', 'steps': 'steps', '测试步骤': 'steps', '步骤': 'steps',
            'expected_result': 'expected_result', '预期结果': 'expected_result',
            'priority': 'priority', '优先级': 'priority',
            'case_type': 'case_type', '用例类型': 'case_type',
            'precondition': 'precondition', '前置条件': 'precondition',
            'case_id': 'case_id', '用例编号': 'case_id',
        }
        cases = []
        for row in reader:
            tc = {'title': '', 'steps': '', 'expected_result': '', 'priority': 'P2', 'case_type': '功能测试', 'module': '', 'precondition': '', 'case_id': ''}
            for key, val in row.items():
                key_lower = key.strip().lower()
                if key_lower in field_map and val:
                    tc[field_map[key_lower]] = val.strip()
            if tc['title']:
                cases.append(tc)
        return cases
